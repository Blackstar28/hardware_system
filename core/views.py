from django.shortcuts import render, redirect
from . import models
from .models import Product, Sale, SaleItem
from .forms import POSFormSet
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse
from django.template.loader import render_to_string
import pdfkit
from django.core.mail import send_mail
from core.utils import send_telegram_message
from django.db.models import F
from .models import PurchaseOrder, PurchaseOrderItem, Product
from .forms import PurchaseOrderForm, POItemFormSet
from .forms import PurchaseOrderForm, PurchaseOrderItemFormSet
from django.db.models import Q
from django.views.generic import ListView
from .models import Delivery
from django.shortcuts import get_object_or_404, redirect
from django.core.paginator import Paginator
from core.utils import notify_low_stock
from django.contrib import messages
from core.models import Product
from django.conf import settings
from django.http import HttpResponse
import requests
from django.http import JsonResponse
from weasyprint import HTML
import tempfile, os
from django.views.decorators.http import require_GET
from decimal import Decimal
from django.utils import timezone
from django.contrib.auth import authenticate, login
import json
from .models import Sale, ArchivedSale
from datetime import datetime
import weasyprint
from .models import ArchivedSale
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook


def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
        
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, "Invalid credentials")

    return render(request, 'core/home.html') 

def purchase_order_list(request):
    return render(request, 'core/purchase_order_list.html', {})

def pos_view(request):
    products = Product.objects.all()
    return render(request, 'pos.html', {'products': products})

def home_view(request):
    return render(request, 'core/home.html')

# Check if user is cashier
def is_cashier(user):
    return user.groups.filter(name='Cashier').exists() or user.is_superuser

# Generate PDF Receipt
# Generate PDF Receipt using WeasyPrint
from decimal import Decimal

def generate_receipt_pdf(request, sale_id):
    try:
        sale = get_object_or_404(Sale, id=sale_id)
        items = SaleItem.objects.filter(sale=sale)

        receipt = []
        gross_total = Decimal("0.00")

        for item in items:
            line_total = item.quantity * item.price
            receipt.append({
                'name': item.product.name,
                'qty': item.quantity,
                'unit': item.price,
                'total': line_total
            })
            gross_total += line_total

        discount_percent = Decimal("10")
        tax_percent = Decimal("12")
        discount_amount = gross_total * (discount_percent / Decimal("100"))
        tax_amount = (gross_total - discount_amount) * (tax_percent / Decimal("100"))
        net_total = (gross_total - discount_amount) + tax_amount

        html_string = render_to_string("core/receipt.html", {
            'sale': sale,
            'receipt': receipt,
            'gross_total': gross_total,
            'discount_percent': int(discount_percent),
            'discount_amount': discount_amount,
            'tax_percent': int(tax_percent),
            'tax_amount': tax_amount,
            'net_total': net_total,
            'receipt_datetime': timezone.localtime(),
        })

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as output:
            HTML(string=html_string).write_pdf(output.name)
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/pdf')
            output.close()
            os.remove(output.name)
            return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f"<h1>❌ PDF Error</h1><pre>{e}</pre>", status=500)



@login_required
def home_view(request):
    return render(request, 'core/home.html')


@user_passes_test(is_cashier)
@login_required
def pos_view(request):
    formset = POSFormSet(request.POST or None)

    if request.method == 'POST':
        try:
            if formset.is_valid():
                sale = Sale.objects.create(cashier=request.user, total=0)
                total = 0
                receipt = []

                for form in formset:
                    if form.cleaned_data and not form.cleaned_data.get('DELETE'):
                        product = form.cleaned_data.get('product')
                        quantity = form.cleaned_data.get('quantity')

                        if not product or not quantity:
                            continue

                        if product.stock >= quantity:
                            SaleItem.objects.create(
                                sale=sale,
                                product=product,
                                quantity=quantity,
                                price=product.price
                            )
                            product.stock -= quantity
                            product.save()
                            notify_low_stock(product)

                            line_total = product.price * quantity
                            total += line_total
                            receipt.append({
                                'name': product.name,
                                'qty': quantity,
                                'unit': product.price,
                                'total': line_total
                            })
                        else:
                            messages.error(request, f"❌ Not enough stock for {product.name}")
                            sale.delete()
                            return redirect('pos_view')

                sale.total = total
                sale.save()

                discount_percent = Decimal("0.10")
                tax_percent = Decimal("0.12")

                discount_amount = total * discount_percent
                taxable_total = total - discount_amount
                tax_amount = taxable_total * tax_percent
                net_total = taxable_total + tax_amount

                messages.success(request, "✅ Sale completed successfully.")
                return redirect('pos_view')
            
            else:
                print("🛑 Formset is invalid:")
                print(formset.errors)
                messages.error(request, "Invalid form data submitted.")
        except Exception as e:
            import traceback
            traceback.print_exc()
            return HttpResponse(f"<h2>🔥 Server Error</h2><pre>{str(e)}</pre>", status=500)

    return render(request, 'core/pos.html', {
        'formset': formset,
        'products': Product.objects.all(),
        'empty_form': formset.empty_form,
    })


def check_and_notify_low_stock():
    low_stock_products = Product.objects.filter(stock__lte=F('low_stock_threshold'))
    if low_stock_products.exists():
        product_list = "\n".join([f"{p.name} (Stock: {p.stock})" for p in low_stock_products])
        send_mail(
            subject='⚠️ Low Stock Alert',
            message=f'The following products are low on stock:\n\n{product_list}',
            from_email='ajkleinf@gmail.com',
            recipient_list=['ajkleinf@gmail.com'],
            fail_silently=False,
        )
    # After saving the sale and reducing stock
check_and_notify_low_stock()

def create_purchase_order(request):
    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST)
        formset = PurchaseOrderItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            po = form.save()
            items = formset.save(commit=False)
            for item in items:
                if item.product and item.quantity:
                    item.purchase_order = po
                    item.save()
            messages.success(request, "✅ Purchase order created successfully.")
            return redirect('create_purchase_order')
    else:
        form = PurchaseOrderForm()
        formset = PurchaseOrderItemFormSet(queryset=PurchaseOrderItem.objects.none())

    return render(request, 'core/create_purchase_order.html', {
        'form': form,
        'formset': formset
    })




def create_purchase_order(request):
    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST)
        formset = POItemFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            purchase_order = form.save()
            items = formset.save(commit=False)

            for item in items:
                item.purchase_order = purchase_order
                item.save()

            return redirect('purchase_order_success')  # You can change this
    else:
        form = PurchaseOrderForm()
        formset = POItemFormSet()

    return render(request, 'core/create_purchase_order.html', {
        'form': form,
        'formset': formset
    })

def create_purchase_order(request):
    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST)
        formset = PurchaseOrderItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            po = form.save()
            items = formset.save(commit=False)
            for item in items:
                item.purchase_order = po
                item.save()
            return redirect('purchase_order_list')  # You can change this later
    else:
        form = PurchaseOrderForm()
        formset = PurchaseOrderItemFormSet(queryset=PurchaseOrderItem.objects.none())
    
    return render(request, 'core/create_purchase_order.html', {
        'form': form,
        'formset': formset
    })
def delivery_list_view(request):
    deliveries = Delivery.objects.all().order_by('-delivery_date')  # newest first
    return render(request, 'core/deliveries.html', {'deliveries': deliveries})

def delivery_list(request):
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    sort_by = request.GET.get('sort', 'delivery_date')

    deliveries = Delivery.objects.all()

    if search_query:
        deliveries = deliveries.filter(
            Q(purchase_order__supplier__name__icontains=search_query) |
            Q(purchase_order__id__icontains=search_query)
        )

    if status_filter == 'delivered':
        deliveries = deliveries.filter(delivered=True)
    elif status_filter == 'not_delivered':
        deliveries = deliveries.filter(delivered=False)

    if sort_by in ['delivery_date', '-delivery_date', 'delivered']:

        deliveries = deliveries.order_by(sort_by)

    # Optional: paginate (10 per page)
    paginator = Paginator(deliveries, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        'deliveries': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'sort_by': sort_by
    }
    return render(request, 'core/delivery_list.html', context)

class DeliveryListView(ListView):
    model = Delivery
    template_name = 'deliveries/delivery_list.html'
    context_object_name = 'deliveries'

    def get_queryset(self):
        queryset = Delivery.objects.select_related('purchase_order', 'purchase_order__supplier')

        search_query = self.request.GET.get('q', '')
        status_filter = self.request.GET.get('status', '')

        if search_query:
            queryset = queryset.filter(
                Q(purchase_order__supplier__name__icontains=search_query) |
                Q(purchase_order__id__icontains=search_query)
            )

        if status_filter == 'delivered':
            queryset = queryset.filter(received=True)
        elif status_filter == 'pending':
            queryset = queryset.filter(received=False)

        sort_by = self.request.GET.get('sort', 'date_received')
        if sort_by in ['date_received', '-date_received']:
            queryset = queryset.order_by(sort_by)

        return queryset
    
def mark_delivery_received(request, pk):
    delivery = get_object_or_404(Delivery, pk=pk)
    delivery.received = True
    delivery.save()
    messages.success(request, f'Delivery {delivery.pk} marked as received.')
    return redirect('delivery_list')

def undo_delivery(request, pk):
    delivery = get_object_or_404(Delivery, pk=pk)
    delivery.received = False
    delivery.save()
    messages.info(request, f'Delivery {delivery.pk} marked as pending.')
    return redirect('delivery_list')

def toggle_delivery_status(request, pk):
    delivery = get_object_or_404(Delivery, pk=pk)
    delivery.delivered = not delivery.delivered
    delivery.save()
    
    if delivery.delivered:
        messages.success(request, "Delivery marked as delivered.")
    else:
        messages.warning(request, "Delivery marked as *not* delivered.")
    
    return redirect('delivery_list')

def pos_sale(request):
    if request.method == "POST":
        product_id = request.POST.get("product_id")
        quantity = int(request.POST.get("quantity", 1))

        try:
            product = Product.objects.get(id=product_id)

            if product.stock >= quantity:
                product.stock -= quantity
                product.save()

                # 🔔 Send alert if stock is low
                notify_low_stock(product)

                messages.success(request, f"{product.name} sold. Remaining stock: {product.stock}")
            else:
                messages.error(request, f"Not enough stock for {product.name}")
        except Product.DoesNotExist:
            messages.error(request, "Product not found.")

        return redirect("pos_sale")  # reloads POS page

    products = Product.objects.all()
    return render(request, "pos.html", {"products": products})

def test_telegram(request):
    message = "✅ Test Telegram message from Django!"

    url = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/sendMessage"
    params = {
        "chat_id": settings.TELEGRAM_CHAT_ID,
        "text": message,
    }

    try:
        response = requests.get(url, params=params)
        response.raise_for_status()
        return HttpResponse("Telegram sent!")
    except Exception as e:
        return HttpResponse(f"Telegram FAILED: {e}")
    
@require_GET
def get_product_price(request):
    product_id = request.GET.get('product_id')
    try:
        product = Product.objects.get(pk=product_id)
        return JsonResponse({'price': product.price})
    except Product.DoesNotExist:
        return JsonResponse({'price': 0})
    
def get_product_price(request):
    product_id = request.GET.get('product_id')
    try:
        product = Product.objects.get(id=product_id)
        return JsonResponse({'price': product.price})
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)


def dashboard_view(request):
    sales = Sale.objects.all()
    gross_total = sum(s.total for s in sales)
    discounted_total = gross_total * Decimal('0.95')
    net_total = discounted_total * Decimal('1.12')

    if request.GET.get("reset") == "true":
        breakdown = [
            {"id": s.id, "total": float(s.total), "cashier": s.cashier.username}
            for s in sales
        ]
        ArchivedSale.objects.create(
            gross_total=gross_total,
            discounted_total=discounted_total,
            net_total=net_total,
            breakdown=json.dumps(breakdown)
        )
        sales.delete()
        return redirect(request.path)

    return render(request, "admin/dashboard.html", {
        "gross_total": gross_total,
        "discounted_total": discounted_total,
        "net_total": net_total,
        "sales": sales,
    })

def download_sales_pdf(request):
    sales = Sale.objects.all()
    gross_total = sum(s.total for s in sales)
    discounted_total = gross_total * Decimal('0.95')
    net_total = discounted_total * Decimal('1.12')

    html_string = render_to_string('admin/sales_pdf.html', {
        'gross_total': gross_total,
        'discounted_total': discounted_total,
        'net_total': net_total,
        'sales': sales
    })

    html = HTML(string=html_string)
    pdf_file = html.write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="today_sales.pdf"'
    return response

def archived_sales_view(request):
    archives = ArchivedSale.objects.all().order_by('-archived_at')
    query = request.GET.get('q')
    date = request.GET.get('date')

    if date:
        try:
            selected_date = datetime.strptime(date, '%Y-%m-%d').date()
            archives = archives.filter(archived_at__date=selected_date)
        except ValueError:
            pass

    results = []
    for archive in archives:
        try:
            breakdown = json.loads(archive.breakdown)
        except json.JSONDecodeError:
            breakdown = []

        if query:
            if not any(query.lower() in str(s.get('cashier', '')).lower() for s in breakdown):
                continue

        archive.parsed_breakdown = breakdown
        results.append(archive)

    return render(request, "admin/archived_sales.html", {
        "archives": results,
        "search_query": query or '',
        "search_date": date or '',
    })

def download_sales_pdf(request):
    from decimal import Decimal
    sales = Sale.objects.all()
    gross_total = sum(s.total for s in sales)
    discounted_total = gross_total * Decimal('0.95')
    net_total = discounted_total * Decimal('1.12')

    html_string = render_to_string('admin/sales_pdf.html', {
        'gross_total': gross_total,
        'discounted_total': discounted_total,
        'net_total': net_total,
        'sales': sales
    })

    html = HTML(string=html_string)
    pdf_file = html.write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="today_sales.pdf"'
    return response




def archived_sales_view(request):
    search_query = request.GET.get("q", "").strip()
    search_date = request.GET.get("date", "").strip()

    archives = ArchivedSale.objects.all().order_by("-archived_at")

    filtered_archives = []
    for archive in archives:
        try:
            breakdown = json.loads(archive.breakdown)
        except json.JSONDecodeError:
            breakdown = []

        archive.parsed_breakdown = breakdown  # attach for display

        # Apply filtering
        matches_query = True
        matches_date = True

        if search_query:
            matches_query = any(search_query.lower() in sale["cashier"].lower() for sale in breakdown)

        if search_date:
            matches_date = str(archive.archived_at.date()) == search_date

        if matches_query and matches_date:
            filtered_archives.append(archive)

    return render(request, "admin/archived_sales_list.html", {
        "archives": filtered_archives,
        "search_query": search_query,
        "search_date": search_date,
    })


def download_archive_pdf(request, archive_id):
    archive = get_object_or_404(ArchivedSale, pk=archive_id)

    try:
        sales = json.loads(archive.breakdown)  # Parse the breakdown field
    except json.JSONDecodeError:
        return HttpResponse("Invalid JSON in ArchivedSale.breakdown")

    template = get_template('admin/archive_pdf.html')
    context = {
        'archive': archive,
        'sales': sales,  # Pass parsed sales data to the template
        'base_path': settings.BASE_DIR,
    }

    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="archive_{archive.id}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF')
    return response

def download_archive_excel(request, archive_id):
    archive = get_object_or_404(ArchivedSale, id=archive_id)
    sales = json.loads(archive.breakdown)  # assuming you have a method to parse breakdown

    wb = Workbook()
    ws = wb.active
    ws.title = "Archived Sales"

    # Add summary info
    ws['A1'] = 'Archived At:'
    ws['B1'] = archive.archived_at.strftime('%B %d, %Y, %I:%M %p')

    ws['A2'] = 'Gross Total:'
    ws['B2'] = f"₱{archive.gross_total:.2f}"

    ws['A3'] = 'Discounted Total:'
    ws['B3'] = f"₱{archive.discounted_total:.2f}"

    ws['A4'] = 'Net Total:'
    ws['B4'] = f"₱{archive.net_total:.2f}"

    # Headers for sales breakdown
    ws.append([])  # empty row
    ws.append(['Sale ID', 'Total', 'Cashier'])

    for sale in sales:
        ws.append([sale['id'], float(sale['total']), sale['cashier']])

    # Serve the file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=archive_{archive.id}.xlsx'
    wb.save(response)
    return response

